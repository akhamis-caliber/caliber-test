import os
import json
import csv
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from fastapi import HTTPException
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import seaborn as sns
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
import base64

class ExportService:
    def __init__(self):
        self.supported_formats = ['pdf', 'excel', 'csv']
        self.default_style = {
            'font_family': 'Arial',
            'font_size': 10,
            'header_font_size': 12,
            'title_font_size': 16
        }

    def export_data(self, data: List[Dict[str, Any]], format_type: str, 
                   report_title: str = "Report", custom_style: Optional[Dict] = None) -> bytes:
        """
        Export data to the specified format
        """
        if format_type not in self.supported_formats:
            raise HTTPException(status_code=400, detail=f"Unsupported format: {format_type}")

        style = {**self.default_style, **(custom_style or {})}

        if format_type == 'pdf':
            return self._export_to_pdf(data, report_title, style)
        elif format_type == 'excel':
            return self._export_to_excel(data, report_title, style)
        elif format_type == 'csv':
            return self._export_to_csv(data, report_title, style)

    def _export_to_pdf(self, data: List[Dict[str, Any]], report_title: str, style: Dict) -> bytes:
        """
        Export data to PDF with charts and formatting
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []

        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=style['title_font_size'],
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=style['header_font_size'],
            spaceAfter=20
        )

        # Add title
        story.append(Paragraph(report_title, title_style))
        story.append(Spacer(1, 20))

        # Add generation timestamp
        timestamp = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(timestamp, styles['Normal']))
        story.append(Spacer(1, 30))

        if data:
            # Create table from data
            if data:
                headers = list(data[0].keys())
                table_data = [headers]  # Header row
                
                for row in data:
                    table_data.append([str(row.get(header, '')) for header in headers])

                # Create table
                table = Table(table_data)
                
                # Style the table
                table_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ])
                table.setStyle(table_style)
                story.append(table)

        # Add summary statistics
        if data:
            story.append(Spacer(1, 30))
            story.append(Paragraph("Summary Statistics", header_style))
            
            # Calculate basic statistics
            numeric_columns = []
            for key in data[0].keys():
                try:
                    float(data[0][key])
                    numeric_columns.append(key)
                except (ValueError, TypeError):
                    continue

            if numeric_columns:
                summary_data = [['Metric', 'Count', 'Average', 'Min', 'Max']]
                for col in numeric_columns:
                    values = [float(row[col]) for row in data if str(row[col]).replace('.', '').isdigit()]
                    if values:
                        summary_data.append([
                            col,
                            len(values),
                            f"{sum(values)/len(values):.2f}",
                            f"{min(values):.2f}",
                            f"{max(values):.2f}"
                        ])

                summary_table = Table(summary_data)
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                ]))
                story.append(summary_table)

        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def _export_to_excel(self, data: List[Dict[str, Any]], report_title: str, style: Dict) -> bytes:
        """
        Export data to Excel with formatting and charts
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Data"

        if not data:
            # Add empty report message
            ws['A1'] = report_title
            ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A4'] = "No data available for export"
        else:
            # Convert data to DataFrame for easier handling
            df = pd.DataFrame(data)
            
            # Add title and timestamp
            ws['A1'] = report_title
            ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Style title
            title_font = Font(size=16, bold=True)
            ws['A1'].font = title_font
            
            # Add data starting from row 4
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws[4]:
                cell.font = header_font
                cell.fill = header_fill

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add summary sheet
            self._add_summary_sheet(wb, df, report_title)

            # Add charts sheet
            self._add_charts_sheet(wb, df, report_title)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _add_summary_sheet(self, wb: Workbook, df: pd.DataFrame, report_title: str):
        """
        Add a summary sheet with statistics
        """
        ws_summary = wb.create_sheet("Summary")
        
        # Add title
        ws_summary['A1'] = f"{report_title} - Summary"
        ws_summary['A1'].font = Font(size=14, bold=True)
        
        # Add timestamp
        ws_summary['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Add basic statistics
        ws_summary['A4'] = "Basic Statistics"
        ws_summary['A4'].font = Font(bold=True)
        
        # Get numeric columns
        numeric_columns = df.select_dtypes(include=['number']).columns
        
        if len(numeric_columns) > 0:
            # Add statistics table
            headers = ['Column', 'Count', 'Mean', 'Std', 'Min', '25%', '50%', '75%', 'Max']
            ws_summary.append(headers)
            
            # Style headers
            for cell in ws_summary[5]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Add statistics for each numeric column
            for i, col in enumerate(numeric_columns, 6):
                stats = df[col].describe()
                ws_summary.append([
                    col,
                    stats['count'],
                    round(stats['mean'], 2),
                    round(stats['std'], 2),
                    round(stats['min'], 2),
                    round(stats['25%'], 2),
                    round(stats['50%'], 2),
                    round(stats['75%'], 2),
                    round(stats['max'], 2)
                ])

    def _add_charts_sheet(self, wb: Workbook, df: pd.DataFrame, report_title: str):
        """
        Add a charts sheet with visualizations
        """
        ws_charts = wb.create_sheet("Charts")
        
        # Add title
        ws_charts['A1'] = f"{report_title} - Charts"
        ws_charts['A1'].font = Font(size=14, bold=True)
        
        # Get numeric columns for charts
        numeric_columns = df.select_dtypes(include=['number']).columns
        
        if len(numeric_columns) > 0:
            # Create a simple bar chart for the first numeric column
            chart_data = df[numeric_columns[0]].value_counts().head(10)
            
            # Add chart data
            ws_charts['A3'] = f"Top 10 {numeric_columns[0]} Values"
            ws_charts['A3'].font = Font(bold=True)
            
            ws_charts['A4'] = "Value"
            ws_charts['B4'] = "Count"
            ws_charts['A4'].font = Font(bold=True)
            ws_charts['B4'].font = Font(bold=True)
            
            for i, (value, count) in enumerate(chart_data.items(), 5):
                ws_charts[f'A{i}'] = value
                ws_charts[f'B{i}'] = count
            
            # Create chart
            chart = BarChart()
            chart.title = f"Distribution of {numeric_columns[0]}"
            chart.x_axis.title = "Value"
            chart.y_axis.title = "Count"
            
            data = Reference(ws_charts, min_col=2, min_row=4, max_row=4+len(chart_data))
            cats = Reference(ws_charts, min_col=1, min_row=5, max_row=4+len(chart_data))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws_charts.add_chart(chart, "D3")

    def _export_to_csv(self, data: List[Dict[str, Any]], report_title: str, style: Dict) -> bytes:
        """
        Export data to CSV format
        """
        buffer = io.StringIO()
        
        if not data:
            # Write empty report message
            buffer.write(f"# {report_title}\n")
            buffer.write(f"# Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            buffer.write("# No data available for export\n")
        else:
            # Write header comment
            buffer.write(f"# {report_title}\n")
            buffer.write(f"# Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            buffer.write("# \n")
            
            # Write data
            fieldnames = list(data[0].keys())
            writer = csv.DictWriter(buffer, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

        csv_content = buffer.getvalue()
        buffer.close()
        
        # Convert to bytes
        return csv_content.encode('utf-8')

    def generate_chart_image(self, data: List[Dict[str, Any]], chart_type: str = 'bar', 
                           title: str = "Chart", x_column: str = None, y_column: str = None) -> str:
        """
        Generate a chart image and return as base64 string
        """
        if not data:
            return ""
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Set up the plot
        plt.figure(figsize=(10, 6))
        plt.style.use('seaborn-v0_8')
        
        if chart_type == 'bar':
            if x_column and y_column:
                plt.bar(df[x_column], df[y_column])
            else:
                # Use first two columns
                cols = list(df.columns)
                if len(cols) >= 2:
                    plt.bar(df[cols[0]], df[cols[1]])
        elif chart_type == 'line':
            if x_column and y_column:
                plt.plot(df[x_column], df[y_column], marker='o')
            else:
                cols = list(df.columns)
                if len(cols) >= 2:
                    plt.plot(df[cols[0]], df[cols[1]], marker='o')
        elif chart_type == 'scatter':
            if x_column and y_column:
                plt.scatter(df[x_column], df[y_column])
            else:
                cols = list(df.columns)
                if len(cols) >= 2:
                    plt.scatter(df[cols[0]], df[cols[1]])
        
        plt.title(title)
        plt.xlabel(x_column or 'X')
        plt.ylabel(y_column or 'Y')
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        # Save to buffer
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        buffer.seek(0)
        
        # Convert to base64
        image_base64 = base64.b64encode(buffer.getvalue()).decode()
        plt.close()
        
        return image_base64

# Create a singleton instance
export_service = ExportService() 